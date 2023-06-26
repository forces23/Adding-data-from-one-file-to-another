from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import win32com.client as win32
import logging

logging.basicConfig(filename='logfile.log', level=logging.DEBUG, format='%(asctime)s %(levelname)s: %(message)s')

# newly created files 
new_files_path="output/OutputExcelFile.xlsx"
new_sheet = "Sheet1"

# ----------------------- Converting XLS -> XLSX ----------------------- 
# could not do this becasue excel was not install locally on the vm that i was using  
def xls_to_xlsx(xls_file_path):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    workbook = excel.Workbooks.Open(xls_file_path)
    
    workbook.SaveAs(xls_file_path+"x", FileFormat = 51)     # FileFormat = 51 -> .xlsx extension
    workbook.Close()                                        # FileFormat = 56 -> .xls extension
    excel.Application.Quit()
    
    xlsx_file_path = xls_file_path+"x"
    
    return xlsx_file_path

# ----------------------- reading xlsx file -----------------------
def read_excel(xlsx_file_path, sheet_name, num_rows):
    logging.info("Reading from xcel file ....")
    print("Reading from xcel file ....")
    
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook[sheet_name]

    data = []
    account_numbers = []
    
    logging.info("     STARTED - Iterating over excel file ... ")
    print("     STARTING - Iterating over excel file now ....")
    
    # iterates through each row until it hits the max which is set to 5
    for row in sheet.iter_rows(values_only=True, max_row=sheet.max_row):
    # for row in sheet.iter_rows(values_only=True, max_row=num_rows):
        data.append(row)    
        
    logging.info("     COMPLETE - Iterating over excel file finished ")
    print("     COMPLETE - Iterating over excel file finished ")
    
    logging.info("     STARTED - Sorting the original excel data now before inserting .... ")
    print("     STARTED - Sorting the original excel data now before inserting .... ")
    
    # sorts the data based on the combine first 3 indexs of each tuble and skips the first indexs specified
    # (change the number how many row you want to skip initially)
    sorted_data = sorted(data[2:], key=lambda x: int(str(x[0] if x[0] is not None else 000) 
                                                     +str(x[1] if x[1] is not None else 00)
                                                     +str(x[2] if x[2] is not None else 0000000000000)) )
    
    logging.info("     COMPLETE - Sorting excel data COMPLETE ") 
    print("     COMPLETE - Sorting excel data COMPLETE ")
    
    # closes the excel sheet that its reading
    workbook.close()
    
    logging.info("Finshed reading xcel file ...")
    print("Finshed reading xcel file ...")
    
    return sorted_data, account_numbers


#  ----------------------- writing to excel file -----------------------
def write_excel(erd_data, foloix_data, account_numbers, num_rows):
    logging.debug("Writing to new xcel file ...")
    print("Writing to new xcel file ...")
    
    workbook = Workbook()  
    sheet = workbook.active  
    sheet.title = new_sheet  
    specified_col_index = 12
    
    # Grabs the actual column letter which should be M
    col_letter = get_column_letter(specified_col_index + 1)
    logging.info("   Inserting at Column Letter: "+col_letter)
    print("   Inserting at Column Letter: "+col_letter)
    sheet.insert_cols(specified_col_index + 1)    
    
    logging.info("     STARTED - Inserting data from original excel into new excel and creating space for new column ...")
    print("     STARTED - Inserting data from original excel into new excel and creating space for new column ...")
    
    # inserts the data from the source xlsx file into the new xlsx file 
    for index,row in enumerate(erd_data, start=1):                
        sheet.append(row)
        # This section shifts all the columns over by 1 index at the specified column 
        # e.g. i am specifying 12 which is column M so that i will have an empty column to work with moving forward
        for col_index in range(22, specified_col_index, -1)  : #22 is the max/last column i am specifiying 
            current_cell = sheet.cell(row=index, column=col_index)
            # logging.info(str(current_cell) +" = "+ str(current_cell.value))
            if str(current_cell.value) == "None":
                sheet.cell(row=index, column=col_index + 1, value="")
            else:
                sheet.cell(row=index, column=col_index + 1, value=current_cell.value)
        # if index == 20:
        #     break
        
    logging.info("     COMPLETE - Finished inserting data from original excel and creating new column ")
    print("     COMPLETE - Finished inserting data from original excel and creating new column ")
    
    logging.info("     STARTED - Checking for matches and inserting data into newly added column now ...")
    print("     STARTED - Checking for matches and inserting data into newly added column now ...")
    
    # Goes through each row and sets the index in the new row to whatever as long as there is a match
    row_index = 1
    for rows in range(1,sheet.max_row + 1):
        logging.debug("row_index : "+ str(row_index))
        
        # Makes sure that the 3 strings that make up the account number has the right amount of leading zeros to match the foliox file 
        branch = str(sheet.cell(row=row_index, column=1).value)
        service_type = str(sheet.cell(row=row_index, column=2).value)
        acc_num = str(sheet.cell(row=row_index, column=3).value)
        # Column 1 check 
        if len(branch) != 3:
            # print("4")
            branch = branch.zfill(3)
        # Column 2 check   
        if service_type is None or len(service_type) != 2:
            # print("5.1")
            service_type = "00"
        elif len(service_type) != 2:
            # print("5.2")
            service_type = service_type.zfill(2)
        # Column 3 check 
        if len(acc_num) != 13:
            # print("6")
            acc_num = acc_num.zfill(13)
            
        account_number = str(branch + service_type + acc_num)
                
        no_match_count = 0
        for index in foloix_data:
            # logging.debug(account_number+ " == " + str(index[0]))
            # print(account_number + " == " + str(index[0]))
            # if row_index == 6:
            #     logging.debug(account_number+ " == " + str(index[0]))
                
            if account_number == str(index[0]):
                # logging.debug("folio data index1 = "+str(index))
                sheet[col_letter + str(row_index)] = str(index[1])
                break
            else:
                no_match_count+=1
                # logging.debug("folio data index1 = "+str(index)+" ** NO MATCH **")
                sheet[col_letter + str(row_index)] = ""
        row_index+=1
        
    logging.info("     COMPLETE - Matches found and data inserted  ")
    print("     COMPLETE - Matches found and data inserted  ")
    
    workbook.save(new_files_path)
    workbook.close()
    
    logging.debug("Finished Writing to new xcel file ...")
    print("Finished Writing to new xcel file ...")


def write_excel_simple(txt_data):
    workbook = Workbook()
    sheet = workbook.active  
    sheet.title = new_sheet 
    
    for row in txt_data:
        sheet.append(row)
        
        
    workbook.save("G:/Bobby/BW-419 file reformat/test/test2.xlsx")
    workbook.close()